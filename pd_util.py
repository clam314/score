from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
import openpyxl
import pandas as pd


def get_tb_after_groupby(table, index, column, method, new_columnName):
    result = table.groupby(index)[column].agg(method).reset_index()
    result.columns = [index, new_columnName]
    return result


def vlookup(out_table, in_table, index):
    return out_table.merge(in_table, on=index, how='left')

def change_sheet_style(fileName,sheetName):
    wb = openpyxl.load_workbook(fileName)
    ws = wb[sheetName]

    title_alig = Alignment(horizontal='center', vertical='center',wrap_text=True)
    title_ft = Font(color='FFFFFF')
    title_fill = PatternFill(fill_type='solid', fgColor='8DB4E2')
    def_alig = Alignment(horizontal='center', vertical='center')
    def_border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin'))

    for column in ws.columns:
        c_tl = column[0]
        c_tl.font = title_ft
        c_tl.fill = title_fill
        c_tl.alignment = title_alig
        c_tl.border = def_border
        title = c_tl.value
        for c in list(column)[1:]:
            c.alignment = def_alig
            c.border = def_border
            __set_number_format(c, title)

    wb.save(fileName)

def __set_number_format(cell, title):
    if title == None : 
        return
    if '评分' in title :
        cell.number_format = '0.00'
    elif '应用ID' == title:
        cell.number_format = '0'
        

